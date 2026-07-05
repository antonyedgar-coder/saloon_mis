import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.utils.dateparse import parse_date


def normalize_header(value):
    return (value or "").strip().lower().replace("_", " ")


def read_tabular_upload(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if filename.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            if not row or all(cell in (None, "") for cell in row):
                continue
            rows.append(["" if cell is None else cell for cell in row])
        workbook.close()
        return rows

    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    return [
        row
        for row in csv.reader(io.StringIO(text))
        if row and any(str(cell).strip() for cell in row)
    ]


def find_column_index(headers, *aliases):
    normalized = [normalize_header(header) for header in headers]
    for alias in aliases:
        alias_norm = normalize_header(alias)
        if alias_norm in normalized:
            return normalized.index(alias_norm)
    return None


def cell_text(row, index):
    if index is None or index >= len(row):
        return ""
    return str(row[index]).strip()


def parse_decimal(value, field_label, row_number):
    text = str(value).strip() if value not in (None, "") else ""
    if not text:
        raise ValueError(f"Row {row_number}: {field_label} is required.")
    try:
        amount = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Row {row_number}: {field_label} must be a number.") from exc
    if amount < 0:
        raise ValueError(f"Row {row_number}: {field_label} cannot be negative.")
    return amount


def parse_required_date(value, field_label, row_number):
    if value in (None, ""):
        raise ValueError(f"Row {row_number}: {field_label} is required.")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    parsed = parse_date(text)
    if parsed:
        return parsed
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Row {row_number}: {field_label} must be YYYY-MM-DD or DD-MM-YYYY."
    )


def csv_template_response(filename, headers, sample_rows=None):
    from django.http import HttpResponse

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in sample_rows or []:
        writer.writerow(row)
    response = HttpResponse(buffer.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
